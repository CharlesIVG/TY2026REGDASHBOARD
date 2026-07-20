#!/usr/bin/env python3
"""
Tokyo Yamathon 2026 - Registration Dashboard
------------------------------------------------------------------
Developed by SxS Partners株式会社 and the project developer, C. Stewart.
Provided to the International Volunteer Group-Japan on loan, for the
Tokyo Yamathon 2026 event only. See NOTICE.md.

(c) SxS Partners株式会社 - All rights reserved.
------------------------------------------------------------------
Derives participant counts from Webscorer registration exports and
writes data/teamsize.json.

WHY THIS EXISTS
The Webscorer JSON API returns Distance, Email, Name and Wave - and
nothing else. There is no roster and no team-size field (confirmed by
the probe run of 19 July 2026). Team size therefore cannot be collected
automatically; it can only come from a manual .xlsx export downloaded
from the Webscorer organiser UI.

PRIVACY - THE LOAD-BEARING RULE
The export contains full names, email addresses, phone numbers and home
addresses for every registrant. This repository is public, and Japan's
APPI and the GDPR both apply to that roster.

  * The .xlsx is read from wherever you point this script. It is never
    copied into the repo.
  * Only aggregate counts are written to data/teamsize.json.
  * Nothing printed to stdout contains a name, an address or an email.

Keep the export outside the repo, or rely on .gitignore (which excludes
*.xlsx) to stop `git add -A` from sweeping it in.

USAGE
    python3 scripts/import_teamsize.py EXPORT.xlsx [MORE.xlsx ...]

    # typical - general and sponsor exports together
    python3 scripts/import_teamsize.py ~/Downloads/*GENERAL*.xlsx \
                                       ~/Downloads/*SPONSOR*.xlsx

Requires openpyxl:  pip3 install openpyxl
"""
import json
import os
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone

try:
    import openpyxl
except ImportError:
    print(
        "ERROR: openpyxl is not installed.\n"
        "       pip3 install openpyxl   (or: pip3 install --break-system-packages openpyxl)",
        file=sys.stderr,
    )
    raise SystemExit(1)

JST = timezone(timedelta(hours=9))

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, "data")
OUT_PATH = os.path.join(DATA_DIR, "teamsize.json")
CONFIG_PATH = os.path.join(DATA_DIR, "webscorer_config.json")

# A member block in the export is a repeated group of
# Last Name / フリガナ / First Name / フリガナ / Email columns. The first
# such block is the team captain; each further block is another member
# who was actually entered. Counting blocks that carry a name is what
# gives team size.
NAME_HEADERS = ("last name", "first name")


def load_course_map() -> dict:
    """Reuse the same course-label mapping the collector uses, so a
    label added there is picked up here too."""
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f).get("courseMap", {})
    except (OSError, json.JSONDecodeError):
        return {
            "quarter": ["half-a-half", "half a half", "halfahalf", "quarter",
                        "10k", "10 km", "ハーフ・ア・ハーフ"],
            "half": ["half", "21k", "21 km", "ハーフ"],
            "full": ["full", "42k", "42 km", "フル"],
        }


def classify_course(value, course_map: dict):
    """Map a Distance string onto full / half / quarter.

    Order is load-bearing: 'half-a-half' contains 'half', so quarter
    patterns must be tested first or every Half-a-Half entry silently
    lands in the Half bucket.
    """
    if not value:
        return None
    v = str(value).strip().lower()
    for course in ("quarter", "half", "full"):
        for pat in course_map.get(course, []):
            if str(pat).lower() in v:
                return course
    return None


def find_name_columns(headers: list) -> list:
    """Indices of every 'Last Name' column. One per member block.

    Matched on header text rather than a fixed position, because the
    general and sponsor exports have different column counts and the
    blocks sit at different offsets in each.
    """
    out = []
    for i, h in enumerate(headers):
        if h and "last name" in str(h).strip().lower():
            out.append(i)
    return out


def find_column(headers: list, *names: str):
    """First column whose header starts with any of `names`."""
    for i, h in enumerate(headers):
        if not h:
            continue
        hl = str(h).strip().lower()
        for n in names:
            if hl.startswith(n.lower()):
                return i
    return None


def read_export(path: str, course_map: dict) -> list:
    """Return one dict per team: {size, course}. No personal data."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    name_cols = find_name_columns(headers)
    dist_col = find_column(headers, "distance", "category")

    if not name_cols:
        raise ValueError(
            f"{os.path.basename(path)}: no 'Last Name' columns found - "
            "is this a Webscorer registration export?"
        )

    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v in (None, "") for v in row):
            continue
        # Team size = number of member blocks that carry a surname.
        size = sum(
            1 for c in name_cols
            if c < len(row) and row[c] not in (None, "")
            and str(row[c]).strip() != ""
        )
        if size == 0:
            continue  # blank or malformed row
        course = classify_course(row[dist_col] if dist_col is not None
                                 and dist_col < len(row) else None, course_map)
        teams.append({"size": size, "course": course})

    return teams


def main() -> int:
    paths = sys.argv[1:]
    if not paths:
        print(__doc__.strip().split("USAGE")[-1].strip(), file=sys.stderr)
        return 2

    missing = [p for p in paths if not os.path.exists(p)]
    if missing:
        for p in missing:
            print(f"ERROR: no such file: {p}", file=sys.stderr)
        return 1

    course_map = load_course_map()

    teams = []
    for p in paths:
        found = read_export(p, course_map)
        # Filename only - never row contents.
        print(f"  {os.path.basename(p)}: {len(found)} teams")
        teams.extend(found)

    if not teams:
        print("ERROR: no teams found in any file.", file=sys.stderr)
        return 1

    people = sum(t["size"] for t in teams)
    dist = Counter(t["size"] for t in teams)

    by_course = {}
    for t in teams:
        c = t["course"]
        if not c:
            continue
        by_course.setdefault(c, Counter())[t["size"]] += 1

    unclassified = sum(1 for t in teams if not t["course"])
    if unclassified:
        print(
            f"  WARNING: {unclassified} team(s) had an unrecognised Distance "
            "value and are counted in the totals but not in byCourse. "
            "Add the label to courseMap in data/webscorer_config.json.",
            file=sys.stderr,
        )

    payload = {
        "_source": (
            "Webscorer registration export. Counts only - no personal data. "
            "Team size is NOT available from the JSON API, so this is a "
            "point-in-time snapshot that only refreshes when a new export "
            "is processed with scripts/import_teamsize.py."
        ),
        "asOf": datetime.now(timezone.utc).astimezone(JST).strftime("%Y-%m-%d"),
        "teams": len(teams),
        "people": people,
        "avgTeamSize": round(people / len(teams), 2),
        "distribution": {str(k): v for k, v in sorted(dist.items())},
        "byCourse": {
            c: {str(k): v for k, v in sorted(counts.items())}
            for c, counts in sorted(by_course.items())
        },
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print()
    print(f"  Teams        {payload['teams']}")
    print(f"  People       {payload['people']}")
    print(f"  Avg per team {payload['avgTeamSize']}")
    print(f"  Sizes        {dict(payload['distribution'])}")
    print()
    print(f"Wrote {os.path.relpath(OUT_PATH, REPO_ROOT)} (as of {payload['asOf']}).")
    print("Commit that file only. Do NOT commit the .xlsx exports.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
